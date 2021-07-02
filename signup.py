from tkinter import *
import openpyxl as xl
from login import loginuh

wb1=xl.load_workbook("book.xlsx")
ws1= wb1.worksheets[0]
ws2=wb1.worksheets[1]

def signupuh():
    #wb1 = xl.load_workbook("book.xlsx")
    #ws1 = wb1.worksheets[0]
    global ws1
    global wb1
    global ws2
    root = Tk()
    root.geometry("600x250")
    root['background']='#99643A'
    #root.configure(bg='brown')
    root.title("Quick Chat")
    def button2_click():
        root.destroy()
        import homepage
    def button1_click():
        #e1.delete(0,END)
        #e1.insert("saved")
        firstname=e1.get()
        lastname=e2.get()
        username=e3.get()
        password=e4.get()
        dob=e5.get()
        disabilities=e6.get()
        
        count = ws2.cell(row=1, column=1).value
        #count=int(count)
        #print(count)
        count=int(count)
        i=1
        unlist=[]
        while(i<=count):
            un=ws1.cell(row=i,column=1).value
            un=str(un)
            unlist.append(un)
            i=i+1
        username=str(username)
        #print(unlist)
        changes=0
        innoruchumma=0
        for i in range(1):
            #print(disabilities)
            if(disabilities.isdigit()==False):
                mylabel=Label(root,text="enter a valid number")
                mylabel.grid(row=12,column=0,columnspan=3,padx=10,pady=10)
                innoruchumma=1
                break;
            correction=int(disabilities)
            #print("correction",correction)
            if(correction<1 or correction>3):
                mylabel=Label(root,text="enter a valid number")
                mylabel.grid(row=12,column=0,columnspan=3,padx=10,pady=10)
                innoruchumma=1
                break;
        
        for i in unlist:
            if(i==username):
                mylabel=Label(root,text="Username exists,change the username and try again")
                mylabel.grid(row=12,column=0,columnspan=3,padx=10,pady=10)
                changes=1
                break;
        if(changes==0 and innoruchumma==0):
            temp=(username,password,firstname,lastname,dob,disabilities)
            ws1.append(temp)
            
            ws2.cell(row=1,column=1).value=count+1
            wb1.save("book.xlsx")
            #mylabel=Label(root,text="saved")
            mylabel=Label(root,text="saved,click next to continue")
            mylabel.grid(row=9,column=0,columnspan=3,padx=10,pady=10)
            button_2 = Button(root, text="Next", padx=20, pady=10, command=button2_click)
            button_2.grid(row=13, column=1)
            #root.destroy()
            
    



    mylabel1=Label(root,text="First Name")
    mylabel1.grid(row=3,column=0)
    e1 = Entry(root, width=35, borderwidth=5)
    e1.grid(row=3, column=1,sticky="ew")
    

    mylabel2=Label(root,text="Last Name")
    mylabel2.grid(row=4,column=0)
    e2 = Entry(root, width=35, borderwidth=5)
    e2.grid(row=4, column=1,sticky="ew")

    mylabel3=Label(root,text="User Name")
    mylabel3.grid(row=5,column=0)
    e3 = Entry(root, width=35, borderwidth=5)
    e3.grid(row=5, column=1,sticky="ew")

    mylabel4=Label(root,text="Password")
    mylabel4.grid(row=6,column=0)
    e4 = Entry(root, width=35, borderwidth=5)
    e4.grid(row=6, column=1,sticky="ew")

    mylabel5=Label(root,text="Birthdate(DD/MM/YYYY)")
    mylabel5.grid(row=7,column=0)
    e5 = Entry(root, width=35, borderwidth=5)
    e5.grid(row=7, column=1,sticky="ew")


    mylabel6=Label(root,text="Are you physically challenged?")
    mylabel6.grid(row=8,column=0,sticky='ew')
    
    
    mylabel7=Label(root,text="1)BLIND\t2)DEAF & DUMB\t3)NONE")
    mylabel7.grid(row=9,column=0,sticky='ew')
    e6 = Entry(root, width=35, borderwidth=5)
    e6.grid(row=8, column=1,sticky="ew")
    

    button_1 = Button(root, text="Ok", padx=20, pady=10, command=button1_click)
    button_1.grid(row=12, column=2,sticky='ew')

    root.mainloop()


