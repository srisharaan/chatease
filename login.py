from tkinter import *
import openpyxl as xl
from utils import *
wb1=xl.load_workbook("book.xlsx")
ws1= wb1.worksheets[0]
ws2=wb1.worksheets[1]

def loginuh():
        #wb1 = xl.load_workbook("book.xlsx")
        #ws1 = wb1.worksheets[0]
    global ws1
    global wb1
    global ws2
    root = Tk()
    root.geometry("600x250")
    root['background']='#99643A'
    #root.configure(bg='brown')
    root.title("hospital management system")
    def button2_click(username):
        #print("usernameuh",username)
        root.destroy()
        utilsuh(username)
        
    def button1_click():
        #e1.delete(0,END)
        #e1.insert("saved")
        username=e3.get()
        username=str(username)
        password=e4.get()
        password=str(password)
        count = ws2.cell(row=1, column=1).value
        #count=int(count)
        #print(count)
        count=int(count)
        i=1
        unlist=[]
        while(i<=count):
            un=ws1.cell(row=i,column=1).value
            pas=ws1.cell(row=i,column=2).value
            if(un==username and password==pas):
                mylabel=Label(root,text="Login success,click next to continue")
                mylabel.grid(row=9,column=0,columnspan=3,padx=10,pady=10)
                button_2 = Button(root, text="Next", padx=20, pady=10, command=button2_click(username))
                button_2.grid(row=11, column=1)
            i=i+1



    mylabel3=Label(root,text="User Name")
    mylabel3.grid(row=5,column=0)
    e3 = Entry(root, width=35, borderwidth=5)
    e3.grid(row=5, column=1,sticky="ew")
    mylabel4=Label(root,text="Password")
    mylabel4.grid(row=6,column=0)
    e4 = Entry(root, width=35, borderwidth=5)
    e4.grid(row=6, column=1,sticky="ew")
    button_1 = Button(root, text="lOGIN", padx=20, pady=10, command=button1_click)
    button_1.grid(row=10, column=1)
    root.mainloop()
loginuh()

