import openpyxl
import os
from tkinter import *
import tkinter
import openpyxl as xl
from stt import hear
from tts import SpeakText
from cam import camuh

def stretuh(myusername,tousername):
    
    myusername=str(myusername)
    tousername=str(tousername)
    path="chats/"

    name=path+myusername+tousername+".xlsx"
    name2=path+tousername+myusername+".xlsx"
    isExist=os.path.exists(name)
    isExist2=os.path.exists(name2)
    if(isExist==False):
        xlsx = openpyxl.Workbook()
        sheet = xlsx.active
        sheet.cell(row=1,column=3).value=1
        xlsx.save(name)
    if(isExist2==False):
        xlsx = openpyxl.Workbook()
        sheet = xlsx.active
        sheet.cell(row=1,column=3).value=2
        xlsx.save(name2)
    wb1=xl.load_workbook(name)
    wb2=xl.load_workbook(name2)
    ws1=wb1.worksheets[0]
    ws2=wb2.worksheets[0]
    wb3=xl.load_workbook("Book.xlsx")
    ws3=wb3.worksheets[0]

    
    def send():
        """Handles sending of messages."""
        msg=entry_field.get()
        ##print(msg)
        summap=myusername+":"+msg
        ##print(summap)
        msg_list.insert(tkinter.END, summap)
        msg=(myusername+":"+msg,0)
        ws1.append(msg)
        ws2.append(msg)
        wb1.save(name)
        wb2.save(name2)
        
        
        
        
    def getprevmsg():
        msg_list.delete(0, tkinter.END)
        
        wb1=xl.load_workbook(name)
        wb2=xl.load_workbook(name2)
        ws1=wb1.worksheets[0]
        ws2=wb2.worksheets[0]
        wb3=xl.load_workbook("Book.xlsx")
        ws3=wb3.worksheets[0]

        tempe=0
        rang=ws3.max_row
        for i in range(1,rang+1):
            usernameuh=ws3['A'+str(i)].value
            if(usernameuh==myusername):
                if(1==int(ws3.cell(row=i,column=6).value)):
                    tempe=1
        if(tempe==1):
            rang=ws1.max_row
            for i in range(1,rang+1):
                msg=ws1['A'+str(i)].value
                ji=ws1['B'+str(i)].value
                ##print(ji)
                if(ji==0):
                    SpeakText(msg)
                    #print(msg)
                    ws1['B'+str(i)].value=1
                    msg_list.insert(tkinter.END, msg)
                    wb1.save(name)
                else:
                    msg_list.insert(tkinter.END, msg)
                    
        else:
        
        #global myusername
        #global tousername
        #global name
        #global name2
        #global ws1
        #global ws2
        
            rang=ws1.max_row
            for i in range(1,rang+1):
                msg=ws1['A'+str(i)].value
                msg_list.insert(tkinter.END, msg)
                wb1.save(name)
    def speak(event=None):
        """Handles sending of messages."""
        msg=hear()
        if(msg==None):
            pass
            #print("speech recognition failed")


        
        msg=myusername+":"+msg
        msg_list.insert(tkinter.END, msg)
        msg=(myusername+":"+msg,0)
        ws1.append(msg)
        ws2.append(msg)
        wb1.save(name)
        wb2.save(name2)

        
    def finspell():
        a=camuh()
        sent_str = ""
        for i in a:
            sent_str += str(i)
        msg=myusername+':'+sent_str
        msg_list.insert(tkinter.END, msg)
        msg=(myusername+":"+msg,0)
        ws1.append(msg)
        ws2.append(msg)
        wb1.save(name)
        wb2.save(name2)

    top = tkinter.Tk()
    top.title("quick chat")

    messages_frame = tkinter.Frame(top)
    my_msg = tkinter.StringVar()  # For the messages to be sent.
    scrollbar = tkinter.Scrollbar(messages_frame)  # To navigate through past messages.
    # Following will contain the messages.
    msg_list = tkinter.Listbox(messages_frame, height=15, width=50, yscrollcommand=scrollbar.set)
    scrollbar.pack(side=tkinter.RIGHT, fill=tkinter.Y)
    msg_list.pack(side=tkinter.LEFT, fill=tkinter.BOTH)
    msg_list.pack()
    messages_frame.pack()
    entry_field = tkinter.Entry(top, textvariable=my_msg)
    entry_field.bind("<Return>", send)
    entry_field.pack()

    
    msg_list.insert(tkinter.END, "quick chat")
    
    send_button = tkinter.Button(top, text="Send", command=send)
    send_button.pack()
    speak_button = tkinter.Button(top, text="Speak", command=speak)
    speak_button.pack()
    
    speak_button = tkinter.Button(top, text="Refresh",command=getprevmsg)
    speak_button.pack()

    speak_button = tkinter.Button(top, text="Finger spell",command=finspell)
    speak_button.pack()

    top.mainloop()
#stretuh("aa","sachin")
