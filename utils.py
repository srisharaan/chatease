import openpyxl  
from tkinter import *
from stret import stretuh
naanu=""
def utilsuh(username):
    global naanu
    naanu=username
    #print(naanu)
    wb = openpyxl.load_workbook('Book.xlsx')  
    ws1= wb.worksheets[0]
    ws2= wb.worksheets[1]
    listuh=[]
    j=2
    temp=ws2.cell(row=1,column=1).value
    for  i in range(temp):
        listuh.append(ws1.cell(row=j,column=1).value)
        j=j+1
    #print(listuh)
    root = Tk() 
    listbox = Listbox(root) 
    listbox.pack(side = LEFT, fill = BOTH) 
    scrollbar = Scrollbar(root) 
    scrollbar.pack(side = RIGHT, fill = BOTH) 
    for values in listuh:
        listbox.insert(END, values) 
	
    listbox.config(yscrollcommand = scrollbar.set) 

    scrollbar.config(command = listbox.yview)
    def selected_item():
        global naanu
        for i in listbox.curselection(): 
            chumma=listbox.get(i)
            #print(chumma)
            #print("naanu",naanu)
            naanu=str(naanu)
            chumma=str(chumma)
            stretuh(naanu,chumma)
            
            root.destroy()
            
        
    btn = Button(root, text='chat', command=selected_item) 
  
    # Placing the button and listbox 
    btn.pack(side='bottom') 
    listbox.pack() 

    root.mainloop() 

